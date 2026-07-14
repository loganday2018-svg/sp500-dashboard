"""
S&P 500 performance updater.
Pulls the latest available market data and writes a workbook with:
  - Daily tab   : most recent trading day's % change (latest close vs prior close)
  - Weekly tab  : rolling last-5-trading-day % change
  - Sector tab  : average daily + weekly change per GICS sector
Dates are auto-detected, so this works any day you run it.
"""
import sys
import os
import subprocess

# When launched headless (pythonw / Task Scheduler) there is no console: sys.stdout
# is None, which breaks prints and any library that writes to it. Redirect to a log
# file so runs are diagnosable and nothing blocks on a missing stream.
LOG_PATH = os.path.join(os.path.expanduser("~"), "sp500_update.log")
if sys.stdout is None or sys.stderr is None:
    _logf = open(LOG_PATH, "a", encoding="utf-8", buffering=1)
    sys.stdout = _logf
    sys.stderr = _logf

# --- make sure required packages exist (auto-install on first run) ---
def _ensure(pkgs):
    import importlib
    missing = []
    for mod, pip_name in pkgs:
        try:
            importlib.import_module(mod)
        except ImportError:
            missing.append(pip_name)
    if missing:
        print(f"Installing required packages: {', '.join(missing)} ...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--quiet", *missing])

_ensure([("pandas", "pandas"), ("openpyxl", "openpyxl"),
         ("requests", "requests"), ("lxml", "lxml")])

import io
import json
from datetime import datetime, timedelta
import requests
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# output next to this script's owner desktop
def _desktop():
    # respects OneDrive / folder redirection via the registry, falls back to ~\Desktop
    try:
        import winreg
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                             r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders")
        val, _ = winreg.QueryValueEx(key, "Desktop")
        val = os.path.expandvars(val)
        if os.path.isdir(val):
            return val
    except Exception:
        pass
    return os.path.join(os.path.expanduser("~"), "Desktop")

DESKTOP = _desktop()
OUT = os.environ.get("SP500_OUT", os.path.join(DESKTOP, "SP500_Weekly_Performance.xlsx"))
DASH_OUT = os.environ.get("SP500_DASH", os.path.join(DESKTOP, "SP500_Dashboard.html"))
TEMPLATE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dashboard_template.html")

# --- 1. constituents ---
print("Fetching S&P 500 constituents...")
html = requests.get("https://en.wikipedia.org/wiki/List_of_S%26P_500_companies",
                    headers={"User-Agent": "Mozilla/5.0"}, timeout=30).text
const = pd.read_html(io.StringIO(html))[0].rename(columns={
    "Symbol": "Ticker", "Security": "Company", "GICS Sector": "Sector"})
const["yf"] = const["Ticker"].str.replace(".", "-", regex=False)
tickers = const["yf"].tolist()
print(f"  {len(tickers)} tickers")

# --- 2. prices via Yahoo chart API ---
# NOTE: we call Yahoo's chart endpoint directly with `requests` instead of using
# yfinance. yfinance's curl_cffi networking layer hangs indefinitely when run under
# Windows Task Scheduler (no interactive console), while plain requests to the same
# Yahoo endpoints works fine there. Each request has a hard timeout, so it can never
# hang, and we fan out across a thread pool for speed.
print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] Downloading latest prices...")
# Yahoo's "spark" endpoint returns many symbols per request, so all ~500 tickers
# come from ~10 batched calls instead of 500 separate ones. This is fast and, just
# as important, avoids the rate-limit stalls and the worker-thread network hang that
# both break other approaches under Windows Task Scheduler. All calls are on the
# main thread with a hard per-request timeout, so a run can never hang.
SPARK = "https://query1.finance.yahoo.com/v7/finance/spark"
UA = {"User-Agent": "Mozilla/5.0"}
SESS = requests.Session()
SESS.headers.update(UA)
BATCH = 20  # Yahoo's spark endpoint rejects requests with more than ~25 symbols

def parse_symbol(entry):
    resp = entry.get("response")
    if not resp:
        return None
    r = resp[0]
    ts = r.get("timestamp")
    cl = r.get("indicators", {}).get("quote", [{}])[0].get("close")
    if not ts or cl is None:
        return None
    idx = pd.to_datetime([pd.Timestamp(t, unit="s").normalize() for t in ts])
    ser = pd.Series(cl, index=idx, dtype="float64").dropna()
    ser = ser[~ser.index.duplicated(keep="last")]
    return ser if not ser.empty else None

series = {}
for start in range(0, len(tickers), BATCH):
    batch = tickers[start:start + BATCH]
    for attempt in range(3):
        try:
            resp = SESS.get(SPARK, params={"symbols": ",".join(batch),
                                           "range": "1mo", "interval": "1d"}, timeout=20)
            resp.raise_for_status()
            for entry in resp.json().get("spark", {}).get("result", []):
                ser = parse_symbol(entry)
                if ser is not None:
                    series[entry["symbol"]] = ser
            break
        except Exception as e:
            if attempt == 2:
                print(f"  batch at {start} failed after 3 tries: {e}")
    print(f"  ...{min(start + BATCH, len(tickers))}/{len(tickers)} fetched")

failed = [t for t in tickers if t not in series]
if len(series) < 400:
    print(f"ERROR: only got {len(series)} tickers, aborting this run.")
    sys.exit(1)
print(f"  got prices for {len(series)} tickers ({len(failed)} missing)")
close = pd.DataFrame(series).sort_index()
close.index = pd.to_datetime(close.index)

if len(close.index) < 2:
    print("ERROR: not enough trading days returned. Try again later.")
    sys.exit(1)

latest_date = close.index[-1]
prior_date = close.index[-2]
week_idx = -6 if len(close.index) >= 6 else 0
week_date = close.index[week_idx]
print(f"  Latest trading day: {latest_date:%Y-%m-%d}")
print(f"  Prior day:          {prior_date:%Y-%m-%d}")
print(f"  Week baseline:      {week_date:%Y-%m-%d}")

# --- 3. build rows ---
rows = []
for _, r in const.iterrows():
    t = r["yf"]
    if t not in close.columns:
        continue
    s = close[t].dropna()
    if len(s) < 2:
        continue
    last_px = float(s.iloc[-1])
    prior_px = float(s.iloc[-2])
    day_pct = (last_px / prior_px - 1) * 100
    wk_window = s.iloc[week_idx:]
    wk_base = float(wk_window.iloc[0])
    wk_pct = (last_px / wk_base - 1) * 100
    rows.append({
        "Ticker": r["Ticker"], "Company": r["Company"], "Sector": r["Sector"],
        "Prior Close": round(prior_px, 2), "Last Close": round(last_px, 2),
        "Day % Change": round(day_pct, 2),
        "Wk-Ago Close": round(wk_base, 2),
        "Week High": round(float(wk_window.max()), 2),
        "Week Low": round(float(wk_window.min()), 2),
        "Week % Change": round(wk_pct, 2),
    })

base = pd.DataFrame(rows)
daily = base[["Ticker", "Company", "Sector", "Prior Close", "Last Close", "Day % Change"]] \
    .sort_values("Day % Change", ascending=False).reset_index(drop=True)
weekly = base[["Ticker", "Company", "Sector", "Wk-Ago Close", "Last Close",
               "Week High", "Week Low", "Week % Change"]] \
    .sort_values("Week % Change", ascending=False).reset_index(drop=True)
sector = base.groupby("Sector").agg(
    **{"Avg Day % Change": ("Day % Change", "mean"),
       "Avg Week % Change": ("Week % Change", "mean")}).round(2) \
    .sort_values("Avg Week % Change", ascending=False).reset_index()
print(f"  {len(base)} companies with data")
stamp = datetime.now().strftime("%Y-%m-%d %H:%M")

# --- 4. build interactive HTML dashboard ---
def build_dashboard():
    if not os.path.exists(TEMPLATE):
        print(f"  (skipping dashboard: template not found at {TEMPLATE})")
        return
    records = base.rename(columns={
        "Ticker": "ticker", "Company": "name", "Sector": "sector",
        "Prior Close": "priorClose", "Last Close": "lastClose", "Day % Change": "dayPct",
        "Wk-Ago Close": "wkAgoClose", "Week High": "weekHigh", "Week Low": "weekLow",
        "Week % Change": "weekPct"}).to_dict(orient="records")
    meta = {
        "updated": stamp,
        "dailyRange": f"{prior_date:%b %d} to {latest_date:%b %d}",
        "weeklyRange": f"{week_date:%b %d} to {latest_date:%b %d}",
        "idxDay": round(float(base["Day % Change"].mean()), 2),
        "idxWeek": round(float(base["Week % Change"].mean()), 2),
    }
    with open(TEMPLATE, "r", encoding="utf-8") as f:
        html = f.read()
    html = html.replace("__DATA__", json.dumps(records)).replace("__META__", json.dumps(meta))
    with open(DASH_OUT, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  Dashboard -> {DASH_OUT}")

build_dashboard()

# --- 5. write formatted workbook ---
blue = PatternFill("solid", fgColor="1F4E78")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, size=14, color="1F4E78")
sub_font = Font(italic=True, size=9, color="808080")
green = Font(color="006100")
red = Font(color="9C0006")
border = Border(bottom=Side(style="thin", color="D9D9D9"))


def style_sheet(ws, ndf, pct_cols, price_cols, title, sub):
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A2"] = sub
    ws["A2"].font = sub_font
    hdr_row = 3
    cols = list(ndf.columns)
    for c, name in enumerate(cols, 1):
        cell = ws.cell(hdr_row, c)
        cell.fill = blue
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for ridx in range(len(ndf)):
        xlrow = hdr_row + 1 + ridx
        for c, name in enumerate(cols, 1):
            cell = ws.cell(xlrow, c)
            cell.border = border
            if name in pct_cols:
                cell.number_format = '0.00"%"'
                cell.font = green if ndf.iloc[ridx][name] >= 0 else red
            elif name in price_cols:
                cell.number_format = '#,##0.00'
    for c, name in enumerate(cols, 1):
        width = max(len(str(name)), *(len(str(v)) for v in ndf.iloc[:, c - 1])) + 3
        ws.column_dimensions[get_column_letter(c)].width = min(width, 42)
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)


# skip the Excel save (only) if the workbook is open; dashboard still updates
xlsx_locked = False
if os.path.exists(OUT):
    try:
        with open(OUT, "a+b"):
            pass
    except PermissionError:
        xlsx_locked = True
        print("\n  NOTE: Excel workbook is open, so the .xlsx was not updated.")
        print("        Close 'SP500_Weekly_Performance.xlsx' to refresh it too.")
        print("        (The dashboard was still refreshed.)")

if not xlsx_locked:
    with pd.ExcelWriter(OUT, engine="openpyxl") as xl:
        daily.to_excel(xl, sheet_name="Daily", index=False, startrow=2)
        weekly.to_excel(xl, sheet_name="Weekly", index=False, startrow=2)
        sector.to_excel(xl, sheet_name="By Sector", index=False, startrow=2)
        wb = xl.book
        style_sheet(wb["Daily"], daily, {"Day % Change"},
                    {"Prior Close", "Last Close"},
                    "S&P 500 - Daily Performance by Company",
                    f"{prior_date:%b %d} close to {latest_date:%b %d} close  |  {len(daily)} companies  |  updated {stamp}")
        style_sheet(wb["Weekly"], weekly, {"Week % Change"},
                    {"Wk-Ago Close", "Last Close", "Week High", "Week Low"},
                    "S&P 500 - Rolling 5-Day Performance by Company",
                    f"{week_date:%b %d} close to {latest_date:%b %d} close  |  {len(weekly)} companies  |  updated {stamp}")
        style_sheet(wb["By Sector"], sector, {"Avg Day % Change", "Avg Week % Change"}, set(),
                    "S&P 500 - Average Performance by Sector",
                    f"Daily ({prior_date:%b %d}->{latest_date:%b %d}) and rolling 5-day  |  updated {stamp}")
    print(f"\nSaved workbook -> {OUT}")

print(f"\nDaily  -> up {(base['Day % Change']>0).sum()}, down {(base['Day % Change']<0).sum()}, avg {base['Day % Change'].mean():.2f}%")
print(f"Weekly -> up {(base['Week % Change']>0).sum()}, down {(base['Week % Change']<0).sum()}, avg {base['Week % Change'].mean():.2f}%")

# open the interactive dashboard (skip with SP500_NO_OPEN=1 for headless runs)
if not os.environ.get("SP500_NO_OPEN"):
    try:
        os.startfile(DASH_OUT)
        print("Opening dashboard...")
    except Exception as e:
        print(f"(Could not auto-open dashboard: {e})")
